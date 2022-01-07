from openpyxl import load_workbook
from openpyxl import Workbook
import ast
kk=[]
cnt=2
rr=0
def intersection(lst1, lst2):
    lst3 = [value for value in lst1 if value in lst2]
    return lst3
workbook = load_workbook(filename="jacc1 im mathoverflow.xlsx")
sheet = workbook.active
for value in sheet.iter_rows(min_row=2, values_only=True):
    #kk.append([])
    for items in list(value):
        kk.append(items)
for i in range(len(kk)):
    if kk[i] != None:
        kk[i] = ast.literal_eval(kk[i])
print(kk)
workbook = Workbook()
sheet = workbook.active
sheet["A1"]="Jaccard"
sheet["B1"]="JNR"
sheet["C1"]="LNR"
for i in range(0,len(kk),2):
    if kk[i+1] != None:
        yy=len(intersection(kk[i], kk[i+1]))
        zz=len(kk[i])+len(kk[i+1])-yy
        jac=yy/zz
        jnr=(len(kk[i+1])-yy)/(len(kk[i+1]))
        lnr=(len(kk[i])-yy)/(len(kk[i]))
        sheet["A"+str(cnt)] = jac
        sheet["B"+str(cnt)] = jnr
        sheet["C"+str(cnt)] = lnr
        cnt=cnt+1
    else:
        jac=0
        jnr=0
        lnr=0
        sheet["A" + str(cnt)] = jac
        sheet["B" + str(cnt)] = jnr
        sheet["C" + str(cnt)] = lnr
        cnt = cnt + 1
workbook.save(filename="JACCJNRLNR.xlsx")


